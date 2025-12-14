"""
Registration Parser

Parses Excel files to extract fund registration data.
"""

import re
from pathlib import Path
from typing import Dict, Optional, Tuple
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .models import FundRegistration
from .file_utils import extract_file_version


def load_registrations(
    registration_path: Path,
    registrations: Dict[str, FundRegistration],
    file_version: Optional[str],
    file_date: Optional[datetime]
) -> Tuple[Optional[str], Optional[datetime]]:
    """
    Load fund registrations from Excel file with temporal data.
    
    Args:
        registration_path: Path to Excel file
        registrations: Dictionary to populate with registrations
        file_version: Current file version (will be updated)
        file_date: Current file date (will be updated)
        
    Returns:
        Tuple of (file_version, file_date)
    """
    if not HAS_OPENPYXL:
        print("[WARNING] openpyxl not available, registration data cannot be loaded")
        return file_version, file_date
    
    if not registration_path.exists():
        print(f"[WARNING] Registration file not found at {registration_path}")
        return file_version, file_date
    
    # Extract version info
    file_version, file_date = extract_file_version(registration_path)
    if file_version:
        print(f"[INFO] Registration file version: {file_version}")
    
    try:
        workbook = openpyxl.load_workbook(registration_path, data_only=True)
        
        if 'Registration' not in workbook.sheetnames:
            print("Warning: 'Registration' sheet not found in Excel file")
            return file_version, file_date
        
        sheet = workbook['Registration']
        
        # Find header row
        headers = []
        header_row = None
        country_columns = {}  # Map country names to column indices
        date_columns = {}  # Map country names to date column indices
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx == 1:  # First row should be headers
                headers = [str(cell).strip() if cell else '' for cell in row]
                
                # Find country columns and associated date columns
                for col_idx, header in enumerate(headers):
                    if not header:
                        continue
                    
                    header_lower = header.lower()
                    
                    # Skip metadata columns
                    if header_lower in ['fund', 'share', 'isin', 'date de clôture', 'unnamed: 0', 'date']:
                        continue
                    
                    # Check if this is a date column (for a country)
                    if 'date' in header_lower or 'expiry' in header_lower or 'expiration' in header_lower:
                        # Try to match to a country (e.g., "Germany Date" -> "Germany")
                        for country_name, col_idx_country in country_columns.items():
                            if country_name.lower() in header_lower:
                                date_columns[country_name] = col_idx
                    else:
                        # Regular country column
                        country_columns[header] = col_idx
                
                header_row = row_idx
                break
        
        if not header_row:
            print("Warning: Could not find header row in registration file")
            return file_version, file_date
        
        # Read data rows
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            fund_name = str(row[0]).strip()
            if not fund_name or fund_name.lower() in ['nan', 'none', '']:
                continue
            
            # Get share class and ISIN
            share_class = str(row[1]).strip() if len(row) > 1 and row[1] else None
            isin = str(row[2]).strip() if len(row) > 2 and row[2] else None
            
            # Create key for this fund/share combination
            if share_class:
                key = f"{fund_name}|{share_class}"
            elif isin:
                key = f"{fund_name}|{isin}"
            else:
                key = fund_name
            
            # Get registered countries with temporal data
            registered_countries = set()
            registration_details = {}
            registration_dates_dict = {}
            expiry_dates_dict = {}
            
            for country, col_idx in country_columns.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value and str(value).strip().upper() in ['R', 'RX', 'Y', 'INSTITUTIONAL', 'RETAIL', 'YES', 'X']:
                        registered_countries.add(country)
                        registration_details[country] = str(value).strip()
                        
                        # Try to get registration date
                        if country in date_columns:
                            date_col_idx = date_columns[country]
                            if date_col_idx < len(row):
                                date_value = row[date_col_idx]
                                if date_value:
                                    # Handle both string dates and datetime objects
                                    if isinstance(date_value, datetime):
                                        registration_dates_dict[country] = date_value.strftime('%Y-%m-%d')
                                    else:
                                        registration_dates_dict[country] = str(date_value).strip()
            
            # Store registration with temporal data
            if key not in registrations:
                registrations[key] = FundRegistration(
                    fund_name=fund_name,
                    share_class=share_class,
                    isin=isin,
                    registered_countries=registered_countries,
                    registration_details=registration_details,
                    registration_dates=registration_dates_dict,
                    expiry_dates=expiry_dates_dict
                )
            else:
                # Merge countries and dates if same fund
                registrations[key].registered_countries.update(registered_countries)
                registrations[key].registration_details.update(registration_details)
                registrations[key].registration_dates.update(registration_dates_dict)
                registrations[key].expiry_dates.update(expiry_dates_dict)
        
        # Calculate statistics
        total_registrations = sum(len(reg.registered_countries) for reg in registrations.values())
        countries_covered = set()
        for reg in registrations.values():
            countries_covered.update(reg.registered_countries)
        
        print(f"[SUCCESS] Loaded {len(registrations)} fund registrations")
        print(f"   Total registrations: {total_registrations}")
        print(f"   Countries covered: {len(countries_covered)}")
    except Exception as e:
        print(f"[ERROR] Error loading registration file: {e}")
        import traceback
        traceback.print_exc()
    
    return file_version, file_date

