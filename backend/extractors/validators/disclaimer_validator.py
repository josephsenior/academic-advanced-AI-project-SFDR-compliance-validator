"""
Disclaimer Validator Module

Validates that required disclaimers are present in documents based on:
- Document content (ESG mentions, performance data, etc.)
- Metadata (client type, management company, new product/strategy)
- Disclaimer glossary (Excel file)

Enhanced with multi-level text matching for improved accuracy.
"""

from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from pydantic import BaseModel, Field

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Import enhanced text matching utilities
from backend.utils.matching.text_matcher import DisclaimerTextMatcher  # type: ignore


class RequiredDisclaimer(BaseModel):
    """Represents a required disclaimer"""
    disclaimer_type: str = Field(description="Type of disclaimer (e.g., 'obam_presentation', 'performance')")
    reason: str = Field(description="Why this disclaimer is required")
    location: Optional[str] = Field(None, description="Where it should appear (e.g., 'slide 1', 'all slides')")
    text: Optional[str] = Field(None, description="Expected disclaimer text")
    language: Optional[str] = Field(None, description="Language of the disclaimer")
    client_type: Optional[str] = Field(None, description="Client type (professional/non_professional)")


class MissingDisclaimer(BaseModel):
    """Represents a missing required disclaimer"""
    disclaimer_type: str
    reason: str
    severity: str = Field(default="error", description="Severity: 'error' or 'warning'")
    location: Optional[str] = None
    expected_text: Optional[str] = None
    match_score: Optional[float] = Field(None, description="Best match score found (0.0-1.0)")
    match_method: Optional[str] = Field(None, description="Matching method used")


class DisclaimerValidationResult(BaseModel):
    """Result of disclaimer validation"""
    required_disclaimers: List[RequiredDisclaimer] = Field(default_factory=list)
    detected_disclaimers: List[str] = Field(default_factory=list, description="Disclaimer types detected in document")
    missing_disclaimers: List[MissingDisclaimer] = Field(default_factory=list)
    present_disclaimers: List[str] = Field(default_factory=list, description="Disclaimer types found in document")
    has_errors: bool = False
    has_warnings: bool = False
    total_required: int = 0
    total_present: int = 0
    total_missing: int = 0


class DisclaimerValidator:
    """
    Validates that required disclaimers are present in documents.
    
    Uses:
    - Document extraction results (content analysis)
    - Metadata (client type, management company, etc.)
    - Disclaimer glossary (Excel file)
    """
    
    def __init__(
        self,
        disclaimer_glossary_path: Optional[str] = None,
        default_glossary_path: str = "dataset/GLOSSAIRE DISCLAIMERS 20231122 .xlsx",
        debug_matching: bool = False
    ):
        """
        Initialize disclaimer validator with enhanced multi-level text matching.
        
        Args:
            disclaimer_glossary_path: Path to disclaimer glossary Excel file
            default_glossary_path: Default path if not provided
            debug_matching: Enable detailed matching logs for troubleshooting
        """
        self.glossary_path = Path(disclaimer_glossary_path or default_glossary_path)
        self.glossary: Dict[str, Dict[str, Dict[str, str]]] = {}
        self.debug_matching = debug_matching
        
        # Initialize enhanced text matcher (always enabled)
        self.text_matcher = DisclaimerTextMatcher(debug=debug_matching)
        
        self._load_glossary()
    
    def _load_glossary(self) -> None:
        """Load disclaimer glossary from Excel file"""
        if not HAS_OPENPYXL:
            print("Warning: openpyxl not available, disclaimer glossary cannot be loaded")
            return
        
        if not self.glossary_path.exists():
            print(f"Warning: Disclaimer glossary not found at {self.glossary_path}")
            return
        
        try:
            workbook = openpyxl.load_workbook(self.glossary_path, data_only=True)
            
            # Load each language sheet
            for sheet_name in ['ENGLISH', 'FRENCH', 'GERMAN']:
                if sheet_name not in workbook.sheetnames:
                    continue
                
                sheet = workbook[sheet_name]
                self.glossary[sheet_name] = {
                    'professional': {},
                    'non_professional': {}
                }
                
                # Find header row
                headers = []
                header_row = None
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if any(cell and 'document' in str(cell).lower() for cell in row if cell):
                        headers = [str(cell).lower() if cell else '' for cell in row]
                        header_row = row_idx
                        break
                
                if not header_row:
                    continue
                
                # Find language column and client type columns
                lang_col_idx = None
                prof_col_idx = None
                non_prof_col_idx = None
                
                for idx, header in enumerate(headers):
                    if 'language' in header or sheet_name.lower() in header:
                        lang_col_idx = idx
                    elif 'professional' in header:
                        prof_col_idx = idx
                    elif 'non' in header and 'professional' in header:
                        non_prof_col_idx = idx
                
                # Read data rows
                for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    doc_type = str(row[0]).strip()
                    if not doc_type or doc_type.lower() in ['nan', 'none', '']:
                        continue
                    
                    # Get text for professional and non-professional
                    if prof_col_idx and row[prof_col_idx]:
                        self.glossary[sheet_name]['professional'][doc_type] = str(row[prof_col_idx]).strip()
                    if non_prof_col_idx and row[non_prof_col_idx]:
                        self.glossary[sheet_name]['non_professional'][doc_type] = str(row[non_prof_col_idx]).strip()
            
            print(f"Loaded disclaimer glossary: {len(self.glossary)} languages")
        except Exception as e:
            print(f"Error loading disclaimer glossary: {e}")
            import traceback
            traceback.print_exc()
    
    def _get_disclaimer_text(
        self,
        disclaimer_type: str,
        language: str = "ENGLISH",
        client_type: str = "non_professional"
    ) -> Optional[str]:
        """Get disclaimer text from glossary"""
        if not self.glossary:
            return None
        
        lang = language.upper()
        if lang not in self.glossary:
            lang = "ENGLISH"
        
        client = "professional" if client_type == "professional" else "non_professional"
        
        if lang in self.glossary and client in self.glossary[lang]:
            return self.glossary[lang][client].get(disclaimer_type)
        
        return None
    
    def _determine_required_disclaimers(
        self,
        extraction_result: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> List[RequiredDisclaimer]:
        """Determine which disclaimers are required based on content and metadata"""
        required = []
        
        # Get metadata
        language = metadata.get('language_code', 'ENGLISH').upper()
        client_type = "professional" if metadata.get('is_professional_client', False) else "non_professional"
        management_company = metadata.get('management_company', '').upper()
        is_new_strategy = metadata.get('is_new_strategy', False)
        is_new_product = metadata.get('is_new_product', False)
        document_type = metadata.get('document_type', 'presentation').lower()
        
        # 1. OBAM Presentation - ALWAYS REQUIRED for presentations
        if 'presentation' in document_type or 'ppt' in document_type:
            obam_text = self._get_disclaimer_text("OBAM Presentation", language, client_type)
            required.append(RequiredDisclaimer(
                disclaimer_type="obam_presentation",
                reason="Mandatory for all presentations",
                location="All slides",
                text=obam_text,
                language=language,
                client_type=client_type
            ))
        
        # 2. Commercial Documentation - Based on management company
        if management_company:
            if 'SAS' in management_company or 'FRANCE' in management_company:
                comm_doc_type = "Commercial documentation : management company = OBAM SAS"
            elif 'GMBH' in management_company or 'GERMANY' in management_company:
                comm_doc_type = "Commercial documentation : management company = OBAM GmbH"
            elif 'LUX' in management_company or 'LUXEMBOURG' in management_company:
                comm_doc_type = "Commercial documentation : management company = OBAM Lux"
            else:
                comm_doc_type = "Commercial documentation"
            
            comm_text = self._get_disclaimer_text(comm_doc_type, language, client_type)
            if comm_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="commercial_documentation",
                    reason=f"Required for management company: {management_company}",
                    location="Document footer or last slide",
                    text=comm_text,
                    language=language,
                    client_type=client_type
                ))
        
        # 3. Content-based disclaimers
        structure = extraction_result.get('structure', {})
        disclaimer_categories = structure.get('disclaimer_categories', {})
        all_categories = set()
        for categories in disclaimer_categories.values():
            all_categories.update(categories)
        
        # Performance disclaimer
        if structure.get('performance_slides') or 'performance' in all_categories:
            perf_text = self._get_disclaimer_text("Performance", language, client_type)
            required.append(RequiredDisclaimer(
                disclaimer_type="performance",
                reason="Performance data is present in the document",
                location="Near performance data",
                text=perf_text,
                language=language,
                client_type=client_type
            ))
        
        # ESG Risk disclaimer
        if 'esg_risk' in all_categories or 'sfdr' in all_categories:
            esg_text = self._get_disclaimer_text("ESG Risk", language, client_type)
            required.append(RequiredDisclaimer(
                disclaimer_type="esg_risk",
                reason="ESG criteria mentioned in document",
                location="Near ESG content",
                text=esg_text,
                language=language,
                client_type=client_type
            ))
        
        # SFDR disclaimers
        full_text = extraction_result.get('full_text', '') or extraction_result.get('text', '')
        full_text_lower = full_text.lower()
        
        if 'article 6' in full_text_lower or 'sfdr article 6' in full_text_lower:
            sfdr6_text = self._get_disclaimer_text("SFDR ART.6", language, client_type)
            if sfdr6_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="sfdr_article_6",
                    reason="SFDR Article 6 mentioned",
                    location="Near SFDR content",
                    text=sfdr6_text,
                    language=language,
                    client_type=client_type
                ))
        
        if 'article 8' in full_text_lower or 'sfdr article 8' in full_text_lower:
            sfdr8_text = self._get_disclaimer_text("SFDR ART.8", language, client_type)
            if sfdr8_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="sfdr_article_8",
                    reason="SFDR Article 8 mentioned",
                    location="Near SFDR content",
                    text=sfdr8_text,
                    language=language,
                    client_type=client_type
                ))
        
        if 'article 9' in full_text_lower or 'sfdr article 9' in full_text_lower:
            sfdr9_text = self._get_disclaimer_text("SFDR ART.9", language, client_type)
            if sfdr9_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="sfdr_article_9",
                    reason="SFDR Article 9 mentioned",
                    location="Near SFDR content",
                    text=sfdr9_text,
                    language=language,
                    client_type=client_type
                ))
        
        # Issuers mentioned
        issuer_mentions = extraction_result.get('issuer_mentions', [])
        if issuer_mentions or 'issuers' in all_categories:
            issuer_text = self._get_disclaimer_text("Issuers mentioned", language, client_type)
            if issuer_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="issuers",
                    reason="Company mentions or logos present",
                    location="Near company mentions",
                    text=issuer_text,
                    language=language,
                    client_type=client_type
                ))
        
        # Opinion disclaimer
        if 'opinion' in all_categories:
            opinion_text = self._get_disclaimer_text("Opinion", language, client_type)
            if opinion_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="opinion",
                    reason="Opinions expressed in document",
                    location="Near opinion content",
                    text=opinion_text,
                    language=language,
                    client_type=client_type
                ))
        
        # Backtest disclaimer
        if 'backtest' in all_categories:
            backtest_text = self._get_disclaimer_text("Back-tested performance", language, client_type)
            if backtest_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="backtest",
                    reason="Back-tested performance data present",
                    location="Near backtest data",
                    text=backtest_text,
                    language=language,
                    client_type=client_type
                ))
        
        # Simulation disclaimer
        if 'simulation' in all_categories:
            sim_text = self._get_disclaimer_text("Simulations of future performance", language, client_type)
            if sim_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="simulation",
                    reason="Future performance simulations present",
                    location="Near simulation content",
                    text=sim_text,
                    language=language,
                    client_type=client_type
                ))
        
        # YtM/YtW disclaimer
        if 'ytm' in all_categories:
            ytm_text = self._get_disclaimer_text("YtM/YtW usage", language, client_type)
            if ytm_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="ytm",
                    reason="YtM or YtW terms mentioned",
                    location="Near YtM/YtW content",
                    text=ytm_text,
                    language=language,
                    client_type=client_type
                ))
        
        # SRI disclaimer
        if 'sri' in all_categories or 'summary risk indicator' in full_text_lower:
            sri_text = self._get_disclaimer_text("SRI", language, client_type)
            if sri_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="sri",
                    reason="Summary Risk Indicator (SRI) mentioned",
                    location="Near SRI content",
                    text=sri_text,
                    language=language,
                    client_type=client_type
                ))
        
        # Money Market Fund
        if 'money_market_fund' in all_categories or 'money market fund' in full_text_lower:
            mmf_text = self._get_disclaimer_text("Money Market Fund", language, client_type)
            if mmf_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="money_market_fund",
                    reason="Money Market Fund mentioned",
                    location="Near MMF content",
                    text=mmf_text,
                    language=language,
                    client_type=client_type
                ))
        
        # New Offer disclaimer
        if is_new_strategy:
            new_offer_text = self._get_disclaimer_text("New offer products (strategy only)", language, client_type)
            if new_offer_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="new_offer",
                    reason="Document references a new strategy",
                    location="Near strategy description",
                    text=new_offer_text,
                    language=language,
                    client_type=client_type
                ))
        
        # RAIF disclaimer (Luxembourg funds)
        if 'raif' in full_text_lower or 'luxembourg funds-raif' in full_text_lower:
            raif_text = self._get_disclaimer_text("Commercial documentation (luxembourg funds-RAIF)", language, client_type)
            if raif_text:
                required.append(RequiredDisclaimer(
                    disclaimer_type="commercial_documentation_raif",
                    reason="RAIF (Luxembourg funds) mentioned",
                    location="Document footer",
                    text=raif_text,
                    language=language,
                    client_type=client_type
                ))
        
        return required
    
    def _detect_present_disclaimers(
        self,
        extraction_result: Dict[str, Any],
        required_disclaimers: List[RequiredDisclaimer]
    ) -> Dict[str, Tuple[bool, float, str]]:
        """
        Detect which disclaimers are present using enhanced multi-level text matching.
        
        Args:
            extraction_result: Document extraction results
            required_disclaimers: List of required disclaimers to check
            
        Returns:
            Dict mapping disclaimer_type to (is_present, confidence, method)
        """
        full_text = extraction_result.get('full_text', '') or extraction_result.get('text', '')
        
        if not full_text:
            return {}
        
        present_disclaimers = {}
        
        # Use enhanced multi-level matching for all disclaimers
        for req in required_disclaimers:
            if req.text:
                # Preprocess disclaimer for caching
                self.text_matcher.preprocess_disclaimer(
                    req.disclaimer_type, 
                    req.text
                )
                
                # Match against document text with multi-level strategy
                is_match, confidence, method = self.text_matcher.match_disclaimer(
                    req.disclaimer_type,
                    req.text,
                    full_text,
                    strict=False
                )
                
                present_disclaimers[req.disclaimer_type] = (is_match, confidence, method)
            else:
                # No text to match, fallback to category-based detection
                is_present = self._check_category_based_detection(extraction_result, req.disclaimer_type)
                present_disclaimers[req.disclaimer_type] = (
                    is_present,
                    1.0 if is_present else 0.0,
                    "category_based"
                )
        
        return present_disclaimers
    
    def _check_category_based_detection(
        self,
        extraction_result: Dict[str, Any],
        disclaimer_type: str
    ) -> bool:
        """
        Fallback category-based detection when no text available.
        
        Args:
            extraction_result: Document extraction results
            disclaimer_type: Type of disclaimer to check
            
        Returns:
            True if category indicators found
        """
        
        structure = extraction_result.get('structure', {})
        disclaimer_categories = structure.get('disclaimer_categories', {})
        
        # Collect all detected categories
        all_categories = set()
        for categories in disclaimer_categories.values():
            all_categories.update(categories)
        
        # Map categories to disclaimer types
        category_to_type = {
            'performance': 'performance',
            'esg_risk': 'esg_risk',
            'issuers': 'issuers',
            'simulation': 'simulation',
            'backtest': 'backtest',
            'new_offer': 'new_offer',
            'ytm': 'ytm',
            'opinion': 'opinion',
            'sfdr': 'sfdr',
            'sfdr_article_6': 'sfdr_article_6',
            'sfdr_article_8': 'sfdr_article_8',
            'sfdr_article_9': 'sfdr_article_9',
            'money_market_fund': 'money_market_fund',
            'sri': 'sri',
        }
        
        # Check if this disclaimer type has category indicators
        if disclaimer_type in category_to_type.values():
            for category in all_categories:
                if category_to_type.get(category) == disclaimer_type:
                    return True
        
        # Special checks for specific types
        full_text = (extraction_result.get('full_text', '') or extraction_result.get('text', '')).lower()
        
        if disclaimer_type == 'obam_presentation':
            return 'oddo bhf asset management' in full_text or 'obam' in full_text
        
        if disclaimer_type == 'commercial_documentation':
            return 'commercial documentation' in full_text or 'documentation commerciale' in full_text
        
        if disclaimer_type == 'commercial_documentation_raif':
            return 'raif' in full_text or 'luxembourg funds-raif' in full_text
        
        return False
    
    def validate(
        self,
        extraction_result: Dict[str, Any],
        metadata: Dict[str, Any],
        document_id: Optional[str] = None
    ) -> DisclaimerValidationResult:
        """
        Validate disclaimers in a document.
        
        Args:
            extraction_result: Document extraction results
            metadata: Document metadata (client type, management company, etc.)
            document_id: Optional document identifier
        
        Returns:
            DisclaimerValidationResult with required, present, and missing disclaimers
        """
        # Determine required disclaimers
        required = self._determine_required_disclaimers(extraction_result, metadata)
        
        # Detect present disclaimers using enhanced matching
        present_with_scores = self._detect_present_disclaimers(extraction_result, required)
        
        # Extract just the present disclaimer types
        present = {
            disc_type 
            for disc_type, (is_present, _, _) in present_with_scores.items() 
            if is_present
        }
        
        # Find missing disclaimers
        required_types = {req.disclaimer_type for req in required}
        missing_types = required_types - present
        
        missing = []
        for req in required:
            if req.disclaimer_type in missing_types:
                # Get the match score even for missing items
                match_info = present_with_scores.get(req.disclaimer_type, (False, 0.0, "not_checked"))
                _, score, method = match_info
                
                missing.append(MissingDisclaimer(
                    disclaimer_type=req.disclaimer_type,
                    reason=req.reason,
                    severity="error",
                    location=req.location,
                    expected_text=req.text,
                    match_score=score,
                    match_method=method
                ))
        
        # Determine if there are errors or warnings
        has_errors = len(missing) > 0
        has_warnings = False  # Can be extended for warnings
        
        return DisclaimerValidationResult(
            required_disclaimers=required,
            detected_disclaimers=list(present),
            missing_disclaimers=missing,
            present_disclaimers=[req.disclaimer_type for req in required if req.disclaimer_type in present],
            has_errors=has_errors,
            has_warnings=has_warnings,
            total_required=len(required),
            total_present=len([r for r in required if r.disclaimer_type in present]),
            total_missing=len(missing)
        )

