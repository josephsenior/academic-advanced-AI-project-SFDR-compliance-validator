"""
Registration of Funds Parser

Parses the Excel file containing fund registration information to validate
country mentions in marketing documents.
"""

import os
from pathlib import Path
from typing import Dict, List, Optional, Set
from pydantic import BaseModel, Field

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class FundRegistration(BaseModel):
    """Represents fund registration information"""
    fund_name: str
    share_class: Optional[str] = None
    isin: Optional[str] = None
    registered_countries: Set[str] = Field(default_factory=set)
    registration_details: Dict[str, str] = Field(default_factory=dict)


class RegistrationParser:
    """
    Parses the Registration of Funds Excel file.
    
    The Excel file contains information about which countries each fund
    is authorized to be commercialized in.
    """
    
    def __init__(
        self,
        registration_file_path: Optional[str] = None,
        default_path: str = "dataset/Registration abroad of Funds_20251008.xlsx"
    ):
        """
        Initialize registration parser.
        
        Args:
            registration_file_path: Path to registration Excel file
            default_path: Default path if not provided
        """
        self.registration_path = Path(registration_file_path or default_path)
        self.registrations: Dict[str, FundRegistration] = {}
        self._load_registrations()
    
    def _load_registrations(self) -> None:
        """Load fund registrations from Excel file"""
        if not HAS_OPENPYXL:
            print("Warning: openpyxl not available, registration data cannot be loaded")
            return
        
        if not self.registration_path.exists():
            print(f"Warning: Registration file not found at {self.registration_path}")
            return
        
        try:
            workbook = openpyxl.load_workbook(self.registration_path, data_only=True)
            
            if 'Registration' not in workbook.sheetnames:
                print("Warning: 'Registration' sheet not found in Excel file")
                return
            
            sheet = workbook['Registration']
            
            # Find header row
            headers = []
            header_row = None
            country_columns = {}  # Map country names to column indices
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_idx == 1:  # First row should be headers
                    headers = [str(cell).strip() if cell else '' for cell in row]
                    
                    # Find country columns (columns with country names)
                    for col_idx, header in enumerate(headers):
                        if header and header.lower() not in ['fund', 'share', 'isin', 'date de clôture', 'unnamed: 0']:
                            # This might be a country column
                            country_columns[header] = col_idx
                    
                    header_row = row_idx
                    break
            
            if not header_row:
                print("Warning: Could not find header row in registration file")
                return
            
            # Read data rows
            current_fund = None
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                fund_name = str(row[0]).strip()
                if not fund_name or fund_name.lower() in ['nan', 'none', '']:
                    continue
                
                # Get share class and ISIN
                share_class = str(row[1]).strip() if len(row) > 1 and row[1] else None
                isin = str(row[2]).strip() if len(row) > 2 and row[2] else None
                
                # Create key for this fund/share combination
                if share_class:
                    key = f"{fund_name}|{share_class}"
                elif isin:
                    key = f"{fund_name}|{isin}"
                else:
                    key = fund_name
                
                # Get registered countries
                registered_countries = set()
                registration_details = {}
                
                for country, col_idx in country_columns.items():
                    if col_idx < len(row):
                        value = row[col_idx]
                        if value and str(value).strip().upper() in ['R', 'RX', 'Y', 'INSTITUTIONAL', 'RETAIL']:
                            registered_countries.add(country)
                            registration_details[country] = str(value).strip()
                
                # Store registration
                if key not in self.registrations:
                    self.registrations[key] = FundRegistration(
                        fund_name=fund_name,
                        share_class=share_class,
                        isin=isin,
                        registered_countries=registered_countries,
                        registration_details=registration_details
                    )
                else:
                    # Merge countries if same fund
                    self.registrations[key].registered_countries.update(registered_countries)
                    self.registrations[key].registration_details.update(registration_details)
            
            print(f"Loaded {len(self.registrations)} fund registrations")
        except Exception as e:
            print(f"Error loading registration file: {e}")
            import traceback
            traceback.print_exc()
    
    def is_registered(
        self,
        fund_name: str,
        country: str,
        share_class: Optional[str] = None,
        isin: Optional[str] = None
    ) -> bool:
        """
        Check if a fund is registered in a specific country.
        
        Args:
            fund_name: Name of the fund
            country: Country name to check
            share_class: Optional share class identifier
            isin: Optional ISIN code
        
        Returns:
            True if registered, False otherwise
        """
        # Try to find registration by various keys
        keys_to_try = []
        
        if share_class:
            keys_to_try.append(f"{fund_name}|{share_class}")
        if isin:
            keys_to_try.append(f"{fund_name}|{isin}")
        keys_to_try.append(fund_name)
        
        for key in keys_to_try:
            if key in self.registrations:
                reg = self.registrations[key]
                # Check exact match first
                if country in reg.registered_countries:
                    return True
                # Check case-insensitive match
                country_lower = country.lower()
                for reg_country in reg.registered_countries:
                    if reg_country.lower() == country_lower:
                        return True
        
        return False
    
    def get_registered_countries(
        self,
        fund_name: str,
        share_class: Optional[str] = None,
        isin: Optional[str] = None
    ) -> Set[str]:
        """
        Get all countries where a fund is registered.
        
        Args:
            fund_name: Name of the fund
            share_class: Optional share class identifier
            isin: Optional ISIN code
        
        Returns:
            Set of registered country names
        """
        keys_to_try = []
        
        if share_class:
            keys_to_try.append(f"{fund_name}|{share_class}")
        if isin:
            keys_to_try.append(f"{fund_name}|{isin}")
        keys_to_try.append(fund_name)
        
        for key in keys_to_try:
            if key in self.registrations:
                return self.registrations[key].registered_countries.copy()
        
        return set()
    
    def validate_country_mentions(
        self,
        mentioned_countries: List[str],
        fund_name: str,
        share_class: Optional[str] = None,
        isin: Optional[str] = None
    ) -> Dict[str, bool]:
        """
        Validate that mentioned countries are registered.
        
        Args:
            mentioned_countries: List of country names mentioned in document
            fund_name: Name of the fund
            share_class: Optional share class identifier
            isin: Optional ISIN code
        
        Returns:
            Dictionary mapping country names to registration status (True/False)
        """
        registered = self.get_registered_countries(fund_name, share_class, isin)
        registered_lower = {c.lower() for c in registered}
        
        validation = {}
        for country in mentioned_countries:
            country_lower = country.lower()
            # Check exact match
            is_registered = country in registered or country_lower in registered_lower
            validation[country] = is_registered
        
        return validation

